#! python3
# Used to get a quick update on Coronavirus stats
import requests
import sys
import bs4
import openpyxl
from openpyxl.utils import get_column_letter


def scrapeWeb():
    url = 'https://bnonews.com/index.php/2020/02/the-latest-coronavirus-cases/'
    res = requests.get(url)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
    tableElem = soup.find_all('table')
    userSearch = sys.argv[1]
    coronaData = {}  # Dictionary to hold the Coronavirus stats

    for table in range(len(tableElem)):
        tableRows = tableElem[table].find_all('tr')
        for tr in tableRows[1:]:
            td = tr.find_all('td')
            row = [i.text for i in td]
            country = row[0]
            confirmedCases = row[1].replace(',', '')
            confirmedDeaths = row[2].replace(',', '')
            notes = row[3]

            # Make sure the key for this country exists:
            coronaData.setdefault(country, {'Cases': '0', 'Deaths': '0', 'Notes': ''})
            coronaData[country]['Cases'] = confirmedCases.strip('*')  # Add number of cases
            coronaData[country]['Deaths'] = confirmedDeaths  # Add number of deaths
            coronaData[country]['Notes'] = notes  # Add any notes that exist
    if userSearch == 'save':
        addSpreadsheet(coronaData)
    else:
        try:
            print('Cases: ' + coronaData[userSearch]['Cases'])
            print('Deaths: ' + coronaData[userSearch]['Deaths'])
            print('Notes: ' + coronaData[userSearch]['Notes'])
        except:
            print('Country not found!')


def addSpreadsheet(statsDictionary):
    wb = openpyxl.load_workbook('coronaTracker.xlsx')
    sheet = wb['Sheet']
    countryList = list(statsDictionary.keys())
    for column in range(2, len(countryList)):
        col_letter = get_column_letter(column)
        country = countryList[column-2]
        sheet[col_letter + '1'] = country
        for row in range(2, len(list(statsDictionary[country].keys()))):
            confirmedCases = statsDictionary[country]['Cases']
            confirmedDeaths = statsDictionary[country]['Deaths']
            confirmedNotes = statsDictionary[country]['Notes']
            sheet.cell(row=row, column=column).value = int(confirmedCases)
            sheet.cell(row=row+1, column=column).value = int(confirmedDeaths)
            sheet.cell(row=row+2, column=column).value = confirmedNotes
    wb.save('coronaTracker.xlsx')


scrapeWeb()
